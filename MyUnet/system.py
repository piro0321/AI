
# -*- coding: utf-8 -*-
import os
import re
import cv2
import glob
import pickle
import keras
import time
import csv
import sys
import openpyxl
import numpy as np
from mlc.function import check_existfile,show,Roi_Reduction,draw_contours,load_image_from_txt,Read_Parameter
from PIL import Image,ImageDraw
from utils.self_unet import unet_2d, unet_2d_GAM
import tensorflow as tf
np.set_printoptions(threshold=400000000)

"""
以下、全体的な処理関数
"""


def load_image_s(filename):
    """
    #   load pickle #
    """
    print("Load Pickle Data.")
    with open(filename,mode='rb') as f:
        image = pickle.load(f)
    print("Load Complete.")
    return image


#Amed Test imageをload 
def test(dict):
    """
    #   #　Amed Test Imageで抽出器の性能を評価する   #
    Parameter
    ---------
    dict    :辞書型
            -Parameter.txtから必要な情報を格納したdict
    Result
    ------
    抽出結果を //aka/share/amed/amed_unet_result　に書き出し
    """
    print("start predict")
    size = dict["size"]
    root = dict["amedroot_path"]

    target_size = (size[0],size[1],1)

    cimage,gimage,data_name= load_image_from_txt(dict)
    #cimage,gimage = load_image_s('./image/image.pkl')

    model = unet_2d_GAM(input_shape=target_size)

    model.load_weights(dict["model_path"])
    os.system("cls")


    for i in range(len(cimage)):

        #start = time.time()
        print("\r{0:d}".format(i),end="")#
        w,h,c = gimage[i].shape
        target = cv2.resize(gimage[i],size)
        target = target.reshape((1,size[0],size[1],1))
        target = target / 255#正規化


        pred = model.predict(target)[0]
        pred = np.reshape(pred, (size[0],size[1],1), order='F')

        #閾値
        
        Threshold = 0.7
        #推測
        pred[pred >= Threshold] = 1
        pred[pred < Threshold] = 0


        pred= np.asarray(pred, dtype=np.uint8)
        val = np.copy(cimage[i])

        """
        val     :検出した領域の輪郭をoriginal画像と重ねた画像
        label   :検出した領域の輪郭
        val     :検出した領域
        """

        pred_resize = cv2.resize(pred,(h,w))
        pred_resize = np.reshape(pred_resize, (w,h,1), order='F')

        val,label,relabel =draw_contours(val,pred_resize)
        
        #pred_resize_2 = np.concatenate((pred_resize,pred_resize),axis=2)
        #pred_resize = np.concatenate((pred_resize_2,pred_resize),axis=2)


        result_image = (relabel/255)*cimage[i]

        #1枚の処理時間
        #elapsed_time = time.time() - start
        #print ("elapsed_time:{0}".format(elapsed_time) + "[sec]")        
        #print("pred_resize.shape",pred_resize.shape)
        #print("cimage[i].shape",cimage[i].shape)
        

        #save result
        result_rootpath = "/amed_unet_result/"+os.path.dirname(data_name[i])
        result_dir_path= root+ result_rootpath

        check_existfile(result_dir_path)
        check_existfile(root+ result_rootpath+"/val")
        check_existfile(root+ result_rootpath+"/target")
        check_existfile(root+ result_rootpath+"/pred")

        cv2.imwrite(root+ result_rootpath+"/val/val_"+os.path.basename(data_name[i]), val)
        cv2.imwrite(root+ "/amed_unet_result/"+data_name[i], result_image*255)
        cv2.imwrite(root+ result_rootpath+"/target/target_"+os.path.basename(data_name[i]), cimage[i])
        cv2.imwrite(root+ result_rootpath+"/pred/"+os.path.basename(data_name[i]), relabel*255)
        #cv2.imwrite(root+ result_rootpath+"/pred/"+os.path.basename(data_name[i]), pred_resize*255)




def unet_result(dict):
    """
    #   推測結果から包含率を求める
    -test.txtのデータと，それに対応した腫瘍領域（ROI）を読み込む
    
    Parameter
    ---------
    dict    :辞書型
            -Parameter.txtを読み込んだdictを受け取る
            -Read_Parameter()を使用する事
    Process
    ------
    包含率（Inclusion rate）を求める
    """
    cimage_list,gimage_list ,imagename_list= [],[],[]
    #root   :"//aka/share/amed"
    data_path = dict["amedabdomen_path"]
    #text_path = dict["amedtest_path"]
    text_path = "./my_testdata.txt"
    amedresult_path = dict["amedresult_path"]
    result_ex,count,count= 0,0,0
    list_result ,list_result= [],[]


    color_cord = (0,255,255)
    thickness = 3
    with open(text_path) as f:
        for i,line in enumerate(f):
            #if i <1600:
            #    continue
            root_path = os.path.basename(line)
           
            #load unet pred and val
            unet_predpath = (amedresult_path+"/"+line+"/pred/*.png").replace('\n', '')
            unet_valpath = (amedresult_path+"/"+line+"/val/*.png").replace('\n', '')

            pred_files = glob.glob(unet_predpath)
            val_files = glob.glob(unet_valpath)



            for fs in pred_files:
                base_path = os.path.basename(fs)
                pred = cv2.imread(fs,cv2.IMREAD_GRAYSCALE)
                
            for fs in val_files:
                val = cv2.imread(fs,1)
            
            #各データまでのパス

            file_path = (data_path+line + "/"+ "*roi.csv").replace('\n', '')
            #file_path = ("//aka/share/amed/amed_abdomen/"+line + "/"+ "*roi.csv").replace('\n', '')


            files = glob.glob(file_path)
            list,list_t1,list_t2 = [],[],[]
            flag = 0
            coord_list = []
            #腫瘍位置座標
            #作るのがめんどくさいため，原始的なつくり
            """
            if lenで長さ図るのは，腫瘍の数分の座標を考慮したから

            例えば 
            腫瘍2つ -> 座標8つ　（腫瘍①->x1 y1 x2 y2 腫瘍②->x3,y3,x4,y4）
            """

            for ff in files:

                with open(ff) as f:
                    """
                    ①ROIの中身読み取り->list型で格納
                    ②改行文字による文字列分断-> split(222\n250) -> 222 , 250
                    """

                    list = (f.read()).split(",")
                    
                    if len(list) > 4:
                        list[3],x3=(list[3]).split("\n")
                        flag = 1


                        if len(list) > 7:
                            list[6],x5=(list[6]).split("\n")

                            if len(list)>10:
                                list[9],x7=(list[9]).split("\n")
                                list[12],_= (list[12]).split("\n")
                                x7,y7,x8,y8 = int(x7),int(list[10]),int(list[11]),int(list[12])
                                flag = 3


                            x5,y5,x6,y6 = int(x5),int(list[7]),int(list[8]),int(list[9])
                            flag = 2
                            

                        x3,y3,x4,y4 = int(x3),int(list[4]),int(list[5]),int(list[6])
                        x1,y1,x2,y2 = int(list[0]),int(list[1]),int(list[2]),int(list[3])

                    else:
                        x1,y1,x2,y2 = int(list[0]),int(list[1]),int(list[2]),int(list[3])
            

            file_path = (data_path +"/" +line + "/"+ "*_denoised.png").replace('\n', '')

            files = glob.glob(file_path)
            for f in files:
                #image_path     :1.2.840.113619.2.256.50119124685.1563429440.1915_denoised.png
                image_path = os.path.basename(f)

                gimage = cv2.imread(f,cv2.IMREAD_GRAYSCALE)
                #gimage = cv2.imread(f,cv2.IMREAD_COLOR)
                

            
            #座標をリストへ
            if flag is 0:
                coord_list.append( ((x1, y1), (x2, y2)))         
                
            if flag is 1:
                coord_list.append( ((x1, y1), (x2, y2)))
                coord_list.append( ((x3, y3), (x4, y4)))

            if flag is 2:
                coord_list.append( ((x1, y1), (x2, y2)))
                coord_list.append( ((x3, y3), (x4, y4)))
                coord_list.append( ((x5, y5), (x6, y6)))

            if flag is 3:     
                coord_list.append( ((x1, y1), (x2, y2)))
                coord_list.append( ((x3, y3), (x4, y4)))
                coord_list.append( ((x5, y5), (x6, y6)))
                coord_list.append( ((x7, y7), (x8, y8)))
                """

                else:
                    val = cv2.rectangle(val, (x1, y1), (x2, y2),color_cord,thickness)
                """

            """
            tumor_coord :全てのroiを映した画像
            val         :書き出した画像
            """
            #腫瘍領域の画像作成
            tumor_coord = np.zeros(pred.shape)
            one_coord = np.zeros(pred.shape)
            tumor_area = 0
            
            for co in range(len(coord_list)):
                #座標受け取り
                one_coord = np.zeros(pred.shape)
                one_area = 0
                coord_1,coord_2 = coord_list[co]
                x1, y1 = coord_1
                x2, y2 = coord_2

                #roi書き出し
                one_coord[y1:y2,x1:x2] = 255
                #tumor_area += np.count_nonzero(one_coord == 255)#画素数の面積                
                one_area = np.count_nonzero(one_coord == 255)#画素数の面積
                #val = cv2.rectangle(gimage, (x1, y1), (x2, y2),(255,0,0),thickness)#図形の書き出し
                #面積チェック
                if one_area > 100000:

                    x1, y1, x2, y2 = Roi_Reduction(( x1, y1, x2, y2),0.6)

                else:
                    x1, y1, x2, y2 = Roi_Reduction(( x1, y1, x2, y2),0.6)

                #全てのチェックが終わったら書き出す
                val = cv2.rectangle(val, (x1, y1), (x2, y2),color_cord,thickness)#図形の書き出し
                cv2.imwrite("./result/new_roi/"+image_path,val)
                tumor_coord[y1:y2,x1:x2] = 255

            
            tumor_area = np.count_nonzero(tumor_coord == 255)
            check_existfile("./unet_result/val/")

            cv2.imwrite("./unet_result/val/"+image_path,val)
            cv2.imwrite("./unet_result/tumor/"+image_path,tumor_coord)
            cv2.imwrite("./unet_result/target/"+image_path,gimage)
            #検出結果の論理和
            result_pred = tumor_coord * pred
            cv2.imwrite("./unet_result/pred/"+image_path,pred)
            cv2.imwrite("./unet_result/result_pred/"+image_path,result_pred)

            result_pred = tumor_coord * pred
            result_area = np.count_nonzero(result_pred >0)

            #大きいroiは0.6倍すると腫瘍の大きさ（決められた範囲）になるらしい
            
            inclusion_rate = result_area /tumor_area

            if inclusion_rate >=1:
                inclusion_rate = 1.0

            list_result.append([i+1,line.replace("\n",""),base_path,inclusion_rate,None])

            result_ex += inclusion_rate
            if inclusion_rate > 0.8:
                count+=1

            print("\r{0:d}".format(i+1),end="")
            #show(tumor_coord)
            #if i ==30:
            #    break
            #os.system("cls")
 

    #包含率の計算結果をxlsxで出力
    ws = openpyxl.Workbook()
    sheet = ws.active
    sheet.title = "Inclusion_rate"
    
    result = result_ex / i
    list_result.append([None,None,None,None,result])
    list_result.append([None,None,None,"80%以上",count])
    sheet.cell(row = 1,column = 1).value = "num"
    sheet.cell(row = 1,column = 2).value = "filename"
    sheet.cell(row = 1,column = 3).value = "dataname"
    sheet.cell(row = 1,column = 4).value = "result"
    sheet.cell(row = 1,column = 5).value = "total"
    for x in range(len(list_result)):
        data = list_result[x]
        for kk in range(4):
            sheet.cell(row = x+2,column = kk+1).value =data[kk]


    ws.save("./result/result_inclusion_rate.xlsx")

if __name__ == "__main__":

    Dict = Read_Parameter("./Parameters.txt")

    #Make_Mytest_Datasettext(Dict)
    test(Dict)    
    unet_result(Dict)




